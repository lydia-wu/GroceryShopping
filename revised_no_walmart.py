import openpyxl

wb = openpyxl.load_workbook('MealCostCalculator.xlsx')

# Items user already has
exclude_items = ['olive oil', 'salt', 'pepper', 'spices', 'garlic', 'ginger', 'sourdough bread']

# Get all ingredients needed
ws_ing = wb['ingredients']
ingredients_needed = []
for row in ws_ing.iter_rows(min_row=2, values_only=True):
    if row[0] and row[2]:
        ingredient_name = row[2].lower() if row[2] else ''
        skip = any(excl in ingredient_name for excl in exclude_items)
        if not skip:
            ingredients_needed.append({
                'code': row[0],
                'meal': row[1],
                'ingredient': row[2],
                'qty_needed': row[3],
                'unit': row[4],
                'total_qty': row[5],
                'cost_total': row[6]
            })

# Target stores (no Walmart)
grocery_stores = ['Costco', 'Safeway', 'H-Mart']
ws_item = wb['ItemizedPurchase']

# Build purchase history database
purchase_db = {}
for store in grocery_stores:
    purchase_db[store] = {}

for row in ws_item.iter_rows(min_row=2, values_only=True):
    date = row[0]
    location = row[2]
    item_name = row[7]
    qty = row[8]
    unit = row[9]
    price = row[10]

    if location in grocery_stores and item_name and price:
        if isinstance(price, str):
            continue
        if isinstance(price, (int, float)) and price > 0:
            item_lower = str(item_name).lower()
            if item_lower not in purchase_db[location]:
                purchase_db[location][item_lower] = []
            purchase_db[location][item_lower].append({
                'date': date,
                'item': item_name,
                'qty': qty,
                'unit': unit,
                'price': price
            })

print("=" * 80)
print("REVISED SHOPPING LIST (NO WALMART)")
print("Prioritizing: Costco for meat, H-Mart for produce")
print("=" * 80)

# Categorize ingredients
meat_items = ['chicken', 'ground turkey', 'turkey']
produce_items = ['kale', 'cucumber', 'dill', 'eggplant', 'tomato', 'carrot', 'celery',
                 'berries', 'apple', 'grapefruit', 'grape', 'lime', 'lemon', 'parsley',
                 'sweet potato', 'squash', 'spinach', 'pomegranate']

# Manually assign based on priorities
assignments = {}

# Manual store assignments (user preferences)
manual_costco_items = ['blueberries', 'berries', 'brami', 'pasta', 'roma tomato', 'tomato',
                       'tomato sauce', 'tomato paste', 'green beans', 'peanut butter', 'avocado']
manual_hmart_items = ['purple potato']
manual_safeway_items = []

for ing in ingredients_needed:
    ing_name = ing['ingredient']
    ing_lower = ing_name.lower()

    # Check manual Costco assignments first
    if any(item in ing_lower for item in manual_costco_items):
        # Force Costco for these items
        found = False
        for item_key in purchase_db['Costco'].keys():
            ing_words = ing_lower.split()
            match_score = sum(1 for word in ing_words if word in item_key and len(word) > 2)
            if match_score > 0:
                purchases = purchase_db['Costco'][item_key]
                if purchases:
                    assignments[ing_name] = {
                        'store': 'Costco',
                        'price': purchases[0]['price'],
                        'item': purchases[0]['item']
                    }
                    found = True
                    break
        if not found:
            # Item not in history, assign to Costco anyway with $0 price
            assignments[ing_name] = {
                'store': 'Costco',
                'price': 0,
                'item': 'To be purchased at Costco (not in history)'
            }
        continue

    # Check manual H-Mart assignments
    if any(item in ing_lower for item in manual_hmart_items):
        # Force H-Mart for these items
        found = False
        for item_key in purchase_db['H-Mart'].keys():
            ing_words = ing_lower.split()
            match_score = sum(1 for word in ing_words if word in item_key and len(word) > 2)
            if match_score > 0:
                purchases = purchase_db['H-Mart'][item_key]
                if purchases:
                    assignments[ing_name] = {
                        'store': 'H-Mart',
                        'price': purchases[0]['price'],
                        'item': purchases[0]['item']
                    }
                    found = True
                    break
        if not found:
            # Item not in history, assign to H-Mart anyway with $0 price
            assignments[ing_name] = {
                'store': 'H-Mart',
                'price': 0,
                'item': 'To be purchased at H-Mart (not in history)'
            }
        continue

    # Check manual Safeway assignments
    if any(item in ing_lower for item in manual_safeway_items):
        # Force Safeway for these items
        found = False
        for item_key in purchase_db['Safeway'].keys():
            ing_words = ing_lower.split()
            match_score = sum(1 for word in ing_words if word in item_key and len(word) > 2)
            if match_score > 0:
                purchases = purchase_db['Safeway'][item_key]
                if purchases:
                    assignments[ing_name] = {
                        'store': 'Safeway',
                        'price': purchases[0]['price'],
                        'item': purchases[0]['item']
                    }
                    found = True
                    break
        if not found:
            # Item not in history, assign to Safeway anyway with $0 price
            assignments[ing_name] = {
                'store': 'Safeway',
                'price': 0,
                'item': 'To be purchased at Safeway (not in history)'
            }
        continue

    # Check if it's meat -> prioritize Costco
    if any(meat in ing_lower for meat in meat_items):
        # Check if Costco has it
        found_at_costco = False
        for item_key in purchase_db['Costco'].keys():
            if any(meat in item_key for meat in meat_items):
                if ing_lower.split()[0] in item_key:
                    found_at_costco = True
                    purchases = purchase_db['Costco'][item_key]
                    if purchases:
                        assignments[ing_name] = {
                            'store': 'Costco',
                            'price': purchases[0]['price'],
                            'item': purchases[0]['item']
                        }
                    break

        if not found_at_costco:
            # Try Safeway for meat
            for item_key in purchase_db['Safeway'].keys():
                if any(meat in item_key for meat in meat_items):
                    if ing_lower.split()[0] in item_key:
                        purchases = purchase_db['Safeway'][item_key]
                        if purchases:
                            assignments[ing_name] = {
                                'store': 'Safeway',
                                'price': purchases[0]['price'],
                                'item': purchases[0]['item']
                            }
                        break

    # Check if it's produce -> prioritize H-Mart
    elif any(prod in ing_lower for prod in produce_items):
        # Try H-Mart first
        found_at_hmart = False
        for item_key in purchase_db['H-Mart'].keys():
            ing_words = ing_lower.split()
            match_score = sum(1 for word in ing_words if word in item_key)
            if match_score > 0:
                found_at_hmart = True
                purchases = purchase_db['H-Mart'][item_key]
                if purchases:
                    assignments[ing_name] = {
                        'store': 'H-Mart',
                        'price': purchases[0]['price'],
                        'item': purchases[0]['item']
                    }
                break

        if not found_at_hmart:
            # Try Costco for produce
            for item_key in purchase_db['Costco'].keys():
                ing_words = ing_lower.split()
                match_score = sum(1 for word in ing_words if word in item_key)
                if match_score > 0:
                    purchases = purchase_db['Costco'][item_key]
                    if purchases:
                        assignments[ing_name] = {
                            'store': 'Costco',
                            'price': purchases[0]['price'],
                            'item': purchases[0]['item']
                        }
                    break

    # For everything else, find best price among remaining stores
    if ing_name not in assignments:
        best_store = None
        best_price = float('inf')
        best_item = None

        for store in grocery_stores:
            for item_key in purchase_db[store].keys():
                ing_words = ing_lower.split()
                match_score = sum(1 for word in ing_words if word in item_key and len(word) > 2)
                if match_score > 0:
                    purchases = purchase_db[store][item_key]
                    if purchases and purchases[0]['price'] < best_price:
                        best_price = purchases[0]['price']
                        best_store = store
                        best_item = purchases[0]['item']

        if best_store:
            assignments[ing_name] = {
                'store': best_store,
                'price': best_price,
                'item': best_item
            }
        else:
            assignments[ing_name] = {
                'store': 'Unknown',
                'price': 0,
                'item': 'Not found in history'
            }

# Print assignments by store
from collections import defaultdict
by_store = defaultdict(list)
store_totals = defaultdict(float)

for ing in ingredients_needed:
    assignment = assignments.get(ing['ingredient'], {'store': 'Unknown', 'price': 0})
    by_store[assignment['store']].append({
        'ingredient': ing['ingredient'],
        'qty': ing['qty_needed'],
        'unit': ing['unit'],
        'price': assignment['price'],
        'meal': ing['code'],
        'item': assignment.get('item', '')
    })
    store_totals[assignment['store']] += assignment['price']

for store in sorted(by_store.keys(), key=lambda s: store_totals[s], reverse=True):
    items = by_store[store]
    total = store_totals[store]
    print(f"\n{'=' * 80}")
    print(f"{store.upper()} (${total:.2f} total, {len(items)} items)")
    print('=' * 80)
    for item in items:
        print(f"  - {item['ingredient']}: {item['qty']} {item['unit']} = ${item['price']:.2f} [Meal {item['meal']}]")
        print(f"    (Historical: {item['item']})")

print(f"\n{'=' * 80}")
print(f"GRAND TOTAL: ${sum(store_totals.values()):.2f}")
print('=' * 80)
