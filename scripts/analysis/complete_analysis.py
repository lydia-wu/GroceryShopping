import openpyxl
from collections import defaultdict

wb = openpyxl.load_workbook('MealCostCalculator.xlsx')

# Get complete ingredients list
print("=" * 80)
print("COMPLETE INGREDIENTS LIST")
print("=" * 80)
ws_ing = wb['ingredients']
ingredients_data = []
for row in ws_ing.iter_rows(min_row=2, values_only=True):
    if row[0]:  # If there's a code
        ingredients_data.append({
            'code': row[0],
            'meal': row[1],
            'ingredient': row[2],
            'qty_needed': row[3],
            'unit': row[4],
            'total_qty': row[5],
            'cost_total': row[6],
            'cost_for_qty': row[7] if isinstance(row[7], (int, float)) else None
        })

# Group by meal
meals = {}
for item in ingredients_data:
    code = item['code']
    if code not in meals:
        meals[code] = {
            'name': item['meal'],
            'ingredients': []
        }
    meals[code]['ingredients'].append(item)

for code in sorted(meals.keys()):
    meal = meals[code]
    print(f"\n{code}. {meal['name']}")
    print("-" * 60)
    for ing in meal['ingredients']:
        print(f"  - {ing['ingredient']}: {ing['qty_needed']} {ing['unit']} (from {ing['total_qty']} {ing['unit']} @ ${ing['cost_total']})")

# Now analyze itemized purchases for grocery stores only
print("\n" + "=" * 80)
print("GROCERY STORE PURCHASE HISTORY")
print("=" * 80)

grocery_stores = ['Costco', 'Safeway', 'H-Mart', 'King Sooper\'s ', 'Trader Joe\'s',
                  'Whole Foods', 'Sprouts', 'Walmart', 'Target']

ws_item = wb['ItemizedPurchase']
purchases = defaultdict(list)

for row in ws_item.iter_rows(min_row=2, values_only=True):
    date = row[0]
    location = row[2]
    item_name = row[7]
    qty = row[8]
    unit = row[9]
    price = row[10]

    if location in grocery_stores and item_name and price:
        # Try to parse price if it's a formula
        if isinstance(price, str):
            # Skip formulas
            continue
        if isinstance(price, (int, float)) and price > 0:
            purchases[location].append({
                'date': date,
                'item': item_name,
                'qty': qty,
                'unit': unit,
                'price': price
            })

for store in grocery_stores:
    if store in purchases:
        print(f"\n{store}: {len(purchases[store])} items")
        # Show recent items (last 10)
        for item in purchases[store][:10]:
            print(f"  {item['date']}: {item['item']} - ${item['price']}")
