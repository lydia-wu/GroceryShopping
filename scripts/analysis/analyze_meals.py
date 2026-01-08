import openpyxl
import json

wb = openpyxl.load_workbook('MealCostCalculator.xlsx')

print("=" * 80)
print("SHEET NAMES:", wb.sheetnames)
print("=" * 80)

# Read ingredients sheet
print("\n=== INGREDIENTS SHEET ===")
if 'ingredients' in wb.sheetnames:
    ws = wb['ingredients']
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0] if rows else []
    print(f"Headers: {headers}")
    print(f"\nTotal rows: {len(rows)}")
    for i, row in enumerate(rows[:50]):  # Print first 50 rows
        print(f"Row {i}: {row}")
else:
    print("Sheet 'ingredients' not found")

# Read totalMealCost sheet
print("\n" + "=" * 80)
print("=== TOTAL MEAL COST SHEET ===")
if 'totalMealCost' in wb.sheetnames:
    ws = wb['totalMealCost']
    rows = list(ws.iter_rows(values_only=True))
    for i, row in enumerate(rows):
        print(f"Row {i}: {row}")
else:
    print("Sheet 'totalMealCost' not found")

# Read itemized sheet - just get the unique locations and items
print("\n" + "=" * 80)
print("=== ITEMIZED PURCHASE SUMMARY ===")
if 'ItemizedPurchase' in wb.sheetnames:
    ws = wb['ItemizedPurchase']
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0] if rows else []
    print(f"Headers: {headers}")
    print(f"\nTotal rows: {len(rows)}")

    # Get unique locations
    locations = set()
    for row in rows[1:]:
        if row[2]:  # Location column
            locations.add(row[2])
    print(f"\nUnique locations: {sorted(locations)}")
else:
    print("Sheet 'ItemizedPurchase' not found")
