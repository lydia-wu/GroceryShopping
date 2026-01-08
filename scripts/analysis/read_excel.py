import openpyxl

wb = openpyxl.load_workbook('MealCostCalculator.xlsx')
print('Sheets:', wb.sheetnames)

for sheet_name in wb.sheetnames:
    print(f'\n=== {sheet_name} ===')
    ws = wb[sheet_name]
    for row in ws.iter_rows(values_only=True):
        print(row)
