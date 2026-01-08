import openpyxl
from collections import defaultdict
from datetime import datetime

wb = openpyxl.load_workbook('MealCostCalculator.xlsx')

# Items user already has
exclude_items = ['olive oil', 'salt', 'pepper', 'spices', 'garlic', 'ginger']

# Get all ingredients needed
ws_ing = wb['ingredients']
ingredients_needed = []
for row in ws_ing.iter_rows(min_row=2, values_only=True):
    if row[0] and row[2]:  # If there's a code and ingredient name
        ingredient_name = row[2].lower() if row[2] else ''
        # Skip if user already has this item
        skip = any(excl in ingredient_name for excl in exclude_items)
        if not skip:
            ingredients_needed.append({
                'code': row[0],
                'meal': row[1],
                'ingredient': row[2],
                'qty_needed': row[3],
                'unit': row[4],
                'total_qty': row[5],
                'cost_total': row[6]
            })

print("=" * 80)
print(f"INGREDIENTS TO PURCHASE ({len(ingredients_needed)} items)")
print("=" * 80)
for item in ingredients_needed:
    print(f"{item['code']}. {item['meal']}: {item['ingredient']} ({item['qty_needed']} {item['unit']})")

# Analyze purchases from grocery stores
grocery_stores = ['Costco', 'Safeway', 'H-Mart', 'Walmart']
ws_item = wb['ItemizedPurchase']

# Build purchase history database
purchase_db = defaultdict(lambda: defaultdict(list))

for row in ws_item.iter_rows(min_row=2, values_only=True):
    date = row[0]
    location = row[2]
    item_name = row[7]
    qty = row[8]
    unit = row[9]
    price = row[10]

    if location in grocery_stores and item_name and price:
        if isinstance(price, str):
            continue
        if isinstance(price, (int, float)) and price > 0:
            # Normalize item name for matching
            item_lower = str(item_name).lower()
            purchase_db[location][item_lower].append({
                'date': date,
                'item': item_name,
                'qty': qty,
                'unit': unit,
                'price': price
            })

print("\n" + "=" * 80)
print("PRICE COMPARISON BY INGREDIENT")
print("=" * 80)

# For each ingredient, find where it was purchased and at what price
ingredient_store_map = {}

for ing in ingredients_needed:
    ing_name_lower = ing['ingredient'].lower()
    ing_name_parts = ing_name_lower.split()

    print(f"\n{ing['ingredient']}:")
    found_prices = {}

    for store in grocery_stores:
        for item_key, purchases in purchase_db[store].items():
            # Try to match ingredient name to purchase history
            # Check if any major word from ingredient appears in purchase item
            match_score = sum(1 for part in ing_name_parts if part in item_key)

            if match_score > 0 or ing_name_lower in item_key:
                # Found a potential match
                if purchases:
                    latest = purchases[0]  # Most recent purchase
                    if store not in found_prices:
                        found_prices[store] = latest
                    print(f"  {store}: ${latest['price']} ({latest['item']})")

    if found_prices:
        # Find cheapest store
        cheapest_store = min(found_prices.keys(), key=lambda s: found_prices[s]['price'])
        ingredient_store_map[ing['ingredient']] = {
            'store': cheapest_store,
            'price': found_prices[cheapest_store]['price'],
            'all_prices': found_prices
        }
        print(f"  → BEST: {cheapest_store} at ${found_prices[cheapest_store]['price']}")
    else:
        print(f"  → No historical data found")
        ingredient_store_map[ing['ingredient']] = {
            'store': 'Unknown',
            'price': 0,
            'all_prices': {}
        }

# Create shopping lists by store
print("\n" + "=" * 80)
print("SHOPPING LIST BY STORE")
print("=" * 80)

shopping_by_store = defaultdict(list)
total_by_store = defaultdict(float)

for ing in ingredients_needed:
    store_info = ingredient_store_map.get(ing['ingredient'], {'store': 'Unknown', 'price': 0})
    store = store_info['store']
    price = store_info['price']

    shopping_by_store[store].append({
        'ingredient': ing['ingredient'],
        'qty': ing['qty_needed'],
        'unit': ing['unit'],
        'price': price,
        'meals': ing['code']
    })
    total_by_store[store] += price

for store in sorted(shopping_by_store.keys(), key=lambda s: total_by_store[s], reverse=True):
    items = shopping_by_store[store]
    total = total_by_store[store]
    print(f"\n{store} (${total:.2f} total, {len(items)} items):")
    for item in items:
        print(f"  - {item['ingredient']}: {item['qty']} {item['unit']} (${item['price']:.2f}) [Meal {item['meals']}]")

print("\n" + "=" * 80)
print(f"GRAND TOTAL: ${sum(total_by_store.values()):.2f}")
print("=" * 80)

# Group ingredients by perishability
print("\n" + "=" * 80)
print("INGREDIENT FRESHNESS CATEGORIES")
print("=" * 80)

highly_perishable = ['kale', 'parsley', 'dill', 'cucumber', 'berries', 'grapefruit', 'eggplant',
                      'chicken', 'spinach', 'tomato', 'celery']
moderately_perishable = ['ground turkey', 'sweet potato', 'carrots', 'pomegranate', 'apple',
                         'squash', 'lemon', 'lime', 'egg', 'yogurt', 'feta', 'cheddar']
shelf_stable = ['canned mackerel', 'pasta', 'tomato sauce', 'tomato paste', 'pearl barley',
                'panko', 'peanut butter', 'meat stock', 'kimchi', 'frozen peas', 'riced cauliflower',
                'pistachio', 'grapes', 'sourdough bread', 'lemon juice', 'parmesan', 'vegan cheddar',
                'red pepper flakes', 'green beans']

print("\nHighly Perishable (use within 3-5 days):")
for ing in ingredients_needed:
    ing_lower = ing['ingredient'].lower()
    if any(p in ing_lower for p in highly_perishable):
        print(f"  - {ing['ingredient']} (Meal {ing['code']})")

print("\nModerately Perishable (use within 7-10 days):")
for ing in ingredients_needed:
    ing_lower = ing['ingredient'].lower()
    if any(p in ing_lower for p in moderately_perishable) and not any(p in ing_lower for p in highly_perishable):
        print(f"  - {ing['ingredient']} (Meal {ing['code']})")

print("\nShelf Stable (can store 2+ weeks):")
for ing in ingredients_needed:
    ing_lower = ing['ingredient'].lower()
    if any(p in ing_lower for p in shelf_stable):
        print(f"  - {ing['ingredient']} (Meal {ing['code']})")
